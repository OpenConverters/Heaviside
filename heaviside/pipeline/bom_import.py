"""Ingest a bare BOM (bill of materials) from a CSV/TSV or Excel (.xlsx) file
into the list-of-dicts shape the cross-reference pipeline consumes.

This lets a user cross-reference a plain component list (e.g. exported from a
PLM / distributor cart / a reference design's BOM tab) without supplying a
whole reference-design PDF. The parser only canonicalises column names and
rows — it never fabricates missing values. ``crossref_pipeline._normalize_bom``
applies further canonicalisation downstream, so this layer is deliberately
forgiving about header spelling but strict about "is there actually data".
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

__all__ = ["BomImportError", "parse_bom_bytes", "parse_bom_file"]


class BomImportError(ValueError):
    """Raised when an uploaded BOM file cannot be parsed into components."""


class _NoPartNumberColumn(BomImportError):
    """The table parsed, rows exist, but no column maps to a part number.

    Internal subclass so the LLM-fallback orchestrator can catch *only* this
    case (a recoverable "messy headers" failure) and retry with an LLM column
    map, while genuinely broken files (empty, no header, no rows) still fail
    fast without burning an LLM call."""


# Map common real-world spreadsheet header spellings → the canonical field
# names the CR pipeline understands (see crossref_pipeline._normalize_bom,
# which additionally maps mpn/part→original_mpn and type/category→component_type).
# Keys are compared lower-cased with surrounding whitespace/punctuation trimmed.
_HEADER_ALIASES: dict[str, str] = {
    "mpn": "original_mpn",
    "part": "original_mpn",
    "part number": "original_mpn",
    "part no": "original_mpn",
    "part#": "original_mpn",
    "manufacturer part number": "original_mpn",
    "manufacturer part no": "original_mpn",
    "manufacturer pn": "original_mpn",
    "mfr part number": "original_mpn",
    "mfr part no": "original_mpn",
    "mfr part #": "original_mpn",
    "mfr pn": "original_mpn",
    "mfg part number": "original_mpn",
    "mfg part no": "original_mpn",
    "mfg pn": "original_mpn",
    "original mpn": "original_mpn",
    "original_mpn": "original_mpn",
    "manufacturer": "manufacturer",
    "mfr": "manufacturer",
    "mfg": "manufacturer",
    "maker": "manufacturer",
    "vendor": "manufacturer",
    "brand": "manufacturer",
    "manufacturer name": "manufacturer",
    "category": "component_type",
    "part category": "component_type",
    "type": "component_type",
    "component type": "component_type",
    "component_type": "component_type",
    "ref": "ref_des",
    "refs": "ref_des",
    "reference": "ref_des",
    "references": "ref_des",
    "ref des": "ref_des",
    "refdes": "ref_des",
    "ref_des": "ref_des",
    "designator": "ref_des",
    "reference designator": "ref_des",
    "value": "value",
    "val": "value",
    "voltage": "rated_voltage",
    "rated voltage": "rated_voltage",
    "voltage rating": "rated_voltage",
    "rated_voltage": "rated_voltage",
    "qty": "quantity",
    "quantity": "quantity",
    "description": "description",
    "description (part)": "description",
    "desc": "description",
    "notes": "notes",
    "note": "notes",
    "comment": "notes",
    "comments": "notes",
}


def _canon_header(raw: str) -> str:
    """Canonicalise a single header cell to a pipeline field name (or a
    lower-cased fallback so unknown columns are still carried through)."""
    key = " ".join(str(raw).strip().lower().replace("_", " ").split())
    if not key:
        return ""

    def _alias(k: str) -> str | None:
        if not k:
            return None
        if k in _HEADER_ALIASES:
            return _HEADER_ALIASES[k]
        # Also try the un-spaced form (e.g. "part#").
        return _HEADER_ALIASES.get(k.replace(" ", ""))

    # 1) Exact match first, so aliases that intentionally carry punctuation
    #    ("part#", "mfr part #") still resolve.
    hit = _alias(key)
    if hit:
        return hit
    # 2) Drop decorative punctuation BOM/quote exporters add — LumiQuote marks
    #    required columns with a trailing "*" ("Offered MPN*"); others use
    #    ":" / "#" / "." / "-". Retry on the cleaned key.
    cleaned = key.strip(" *:#.-")
    hit = _alias(cleaned)
    if hit:
        return hit
    # 3) Quote/ERP exporters prefix the real column with a qualifier word
    #    ("Offered MPN", "Supplier MPN"). Retry without a leading qualifier so
    #    the underlying alias ("mpn", "manufacturer", …) still resolves.
    for prefix in ("offered ", "supplier ", "vendor "):
        if cleaned.startswith(prefix):
            hit = _alias(cleaned[len(prefix):].strip())
            if hit:
                return hit
    return (cleaned or key).replace(" ", "_")


def _rows_to_components(
    headers: list[str],
    rows: list[list[Any]],
    *,
    header_overrides: dict[str, str] | None = None,
) -> list[dict[str, Any]]:
    """Map a header row + data rows into canonical component dicts.

    ``header_overrides`` maps an exact source-header string → canonical field
    name; it wins over the deterministic alias table. It is supplied by the
    LLM column-mapper fallback (which only *identifies* columns — the values
    still come verbatim from the cells, never fabricated)."""
    overrides = _normalize_overrides(header_overrides) if header_overrides else {}
    canon = [overrides.get(_norm_key(h)) or _canon_header(h) for h in headers]
    if not any(canon):
        raise BomImportError("BOM file has no usable header row.")
    components: list[dict[str, Any]] = []
    for raw_row in rows:
        row: dict[str, Any] = {}
        for i, field in enumerate(canon):
            if not field or i >= len(raw_row):
                continue
            cell = raw_row[i]
            if cell is None:
                continue
            text = str(cell).strip()
            if text == "":
                continue
            # Don't clobber a real value with a later duplicate-mapped blank.
            row.setdefault(field, text)
        if row:
            components.append(row)
    if not components:
        raise BomImportError("BOM file parsed but contained no component rows (all rows empty).")
    if not any("original_mpn" in c for c in components):
        raise _NoPartNumberColumn(
            "BOM file has no recognisable part-number column. Include a column "
            "named one of: MPN, Part, Part Number, or Manufacturer Part Number."
        )
    return components


def _norm_key(raw: Any) -> str:
    """Whitespace/case-insensitive key for matching an override header to an
    actual header cell (BOM exports pad headers with stray spaces, e.g.
    `" MFG_PN"`)."""
    return " ".join(str(raw).strip().lower().split())


def _normalize_overrides(overrides: dict[str, str]) -> dict[str, str]:
    """Index an LLM ``{canonical_field: source_header}`` map as
    ``{normalized_source_header: canonical_field}`` for matching against the
    real header row. Only canonical fields the pipeline understands are kept;
    anything else (or a null/blank source) is dropped (no fabrication)."""
    out: dict[str, str] = {}
    valid_fields = set(_HEADER_ALIASES.values())
    for field, source_header in overrides.items():
        if field not in valid_fields:
            continue
        if not isinstance(source_header, str) or not source_header.strip():
            continue
        out[_norm_key(source_header)] = field
    return out


def _read_csv_table(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    """Decode + sniff a CSV/TSV into (headers, data_rows). No column mapping."""
    text = raw.decode("utf-8-sig", errors="replace")
    if not text.strip():
        raise BomImportError("CSV file is empty.")
    # Sniff the delimiter from the first non-trivial chunk; fall back to comma.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    table = [row for row in reader]
    # Drop leading fully-blank lines, then take the first row as the header.
    table = [r for r in table if any(str(c).strip() for c in r)]
    if not table:
        raise BomImportError("CSV file has no rows.")
    headers, *rows = table
    return headers, rows


def _read_xlsx_table(raw: bytes) -> tuple[list[str], list[list[Any]]]:
    """Read the active sheet of an .xlsx into (headers, data_rows)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency present in env
        raise BomImportError(
            "openpyxl is required to read .xlsx files; convert to CSV instead."
        ) from exc
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise BomImportError(f"could not open Excel workbook: {exc}") from exc
    try:
        ws = wb.active
        table = [
            list(r)
            for r in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() for c in r)
        ]
    finally:
        wb.close()
    if not table:
        raise BomImportError("Excel sheet has no rows.")
    headers, *rows = table
    return [str(h) if h is not None else "" for h in headers], rows


def _llm_available() -> bool:
    """True if an LLM API key is configured (else the mapper is skipped)."""
    import os

    return bool(os.environ.get("MOONSHOT_API_KEY") or os.environ.get("OPENAI_API_KEY"))


def _llm_header_overrides(headers: list[str], rows: list[list[Any]]) -> dict[str, str]:
    """Ask the bom-header-mapper agent which columns map to which canonical
    fields, returning a ``{source_header: canonical_field}`` override map.

    The LLM only *names columns* — it never reads or fabricates values. This is
    BEST-EFFORT: any failure (no key, agent error, bad output) returns ``{}`` and
    is logged, so a parseable BOM still goes through deterministic aliasing. The
    "no part-number column" error is raised downstream by ``_rows_to_components``,
    not here — so it fires whether or not the LLM ran."""
    import json

    from heaviside.agents.llm_call import LLMCallError, call_agent_json

    if not _llm_available():
        return {}
    # Feed the header row + a few sample data rows so the agent can tell a real
    # manufacturer MPN column apart from an internal/house part-number column,
    # and recognise non-obvious column names (e.g. LOCATION = ref designator).
    sample = [
        {str(headers[i]): ("" if i >= len(r) or r[i] is None else str(r[i]))
         for i in range(len(headers))}
        for r in rows[:5]
    ]
    user_message = (
        "Header row (exact strings):\n"
        + json.dumps(list(headers), ensure_ascii=False)
        + "\n\nSample data rows:\n"
        + json.dumps(sample, ensure_ascii=False)
    )
    try:
        result = call_agent_json("bom-header-mapper", user_message, json_mode=True)
    except LLMCallError as exc:
        logger.warning("bom-header-mapper unavailable, using deterministic aliasing: %s", exc)
        return {}
    if not isinstance(result, dict):
        return {}
    # Drop the rationale / any null entries; keep only canonical→header strings.
    return {
        k: v for k, v in result.items()
        if k != "rationale" and isinstance(v, str) and v.strip()
    }


def parse_bom_bytes(
    raw: bytes, filename: str, *, allow_llm: bool = True
) -> list[dict[str, Any]]:
    """Parse raw file bytes into a BOM component list, choosing CSV vs XLSX
    by the filename extension. Raises :class:`BomImportError` on any failure
    (unsupported type, empty file, no header, no rows, no part-number column).

    When ``allow_llm`` is set (and an API key is configured), the LLM
    column-mapper ALWAYS runs and identifies which column is each canonical
    field (manufacturer part number, ref designator, value, category, …) —
    catching non-standard headers (e.g. ``LOCATION`` = ref designator, ``MFG_PN``
    = MPN) that the deterministic alias table doesn't know. Its mapping wins per
    column; unmapped columns fall back to deterministic aliasing. The LLM only
    selects among the file's own columns — it never invents values, so every
    emitted value still comes verbatim from a real cell. The mapper is
    best-effort: if it's unavailable, deterministic aliasing alone is used."""
    if not raw:
        raise BomImportError("uploaded file is empty.")
    name = (filename or "").lower()
    if name.endswith(".xls"):
        raise BomImportError("legacy .xls is not supported — re-save as .xlsx or export to CSV.")
    if name.endswith((".xlsx", ".xlsm")):
        headers, rows = _read_xlsx_table(raw)
    elif name.endswith((".csv", ".tsv", ".txt", "")):
        headers, rows = _read_csv_table(raw)
    else:
        # Unknown extension: try CSV (most BOM exports are text) before giving up.
        try:
            headers, rows = _read_csv_table(raw)
        except BomImportError as exc:
            raise BomImportError(
                f"unsupported BOM file type {filename!r}; use .csv or .xlsx ({exc})"
            ) from exc

    # The LLM column-mapper runs on every parse (best-effort); its per-column
    # mapping wins, deterministic aliasing fills the rest. _rows_to_components
    # raises _NoPartNumberColumn if NEITHER finds a part-number column.
    overrides = _llm_header_overrides(headers, rows) if allow_llm else {}
    return _rows_to_components(headers, rows, header_overrides=overrides)


# Convenience alias used by the API layer.
parse_bom_file = parse_bom_bytes
